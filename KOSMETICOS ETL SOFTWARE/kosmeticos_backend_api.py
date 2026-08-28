import os
import re
import random
import shutil
import logging
from datetime import date, timedelta
import webbrowser
from typing import Optional, Tuple, Dict, Any, List

import pandas as pd
from openpyxl import load_workbook
from fastapi import FastAPI, File, UploadFile, HTTPException, Query, status
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
import uvicorn

# Configuración de Logging
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger("KosmeticosETL")

# Constantes de Entorno
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_DIR = os.path.join(BASE_DIR, "base_de_datos")
WEB_DIR = os.path.join(BASE_DIR, "web")
PLANTILLAS_DIR = os.path.join(BASE_DIR, "Plantillas")
DICCIONARIOS_DIR = os.path.join(BASE_DIR, "diccionarios")

for folder in [DB_DIR, WEB_DIR, PLANTILLAS_DIR, DICCIONARIOS_DIR]:
    os.makedirs(folder, exist_ok=True)

COLUMNAS_OBLIGATORIAS_MAIN = [
    "SKU_Padre", "SKU_Variante", "Marca", "Nombre_General",
    "Titulo_MercadoLibre", "Precio_Regular", "Stock"
]

MAPEO_CATEGORIAS_ML = {
    "Cuidado Facial": [
        "crema", "cream", "moisturizer", "moisturising", "calming cream", "gel", "lotion", "locion",
        "serum", "sérum", "ampoule", "ampolla", "esencia", "essence", "pdrn", "exosomas", "treatment",
        "tónico", "tonico", "toner", "bruma", "mist", "pad", "pads", "discos",
        "exfoliante", "limpiador", "cleanser", "cleansing", "peeling", "gommage", "scrub", 
        "gel de limpieza", "espuma limpiadora", "desmaquillante", "aceite limpiador", "wash",
        "bloqueador", "protector solar", "sunscreen", "sun cream", "sun stick", "sunblock",
        "mascarilla", "sheet mask", "mask", "cuidado facial", "facial", "skincare"
    ],
    "Maquillaje": [
        "rubor", "rubores", "blush", "cheek", "delineador", "eyeliner", "pen liner", "pencil liner",
        "tinta de ojos", "sombra", "eyeshadow", "paleta de sombras", "shadow", "mascara de pestañas",
        "máscara de pestañas", "rimel", "mascara", "lash", "bb cream", "cc cream", "base de maquillaje",
        "foundation", "tone up", "cushion", "corrector", "concealer", "labial", "tint", "tinta de labios",
        "gloss", "bálsamo labial", "balsamo labial", "lip balm", "lipstick"
    ],
    "Cuidado Corporal": [
        "loción corporal", "crema corporal", "jabón corporal", "body wash", 
        "body lotion", "body cream", "body scrub", "corporal", "body"
    ],
    "Cuidado Capilar": [
        "shampoo", "champú", "acondicionador", "hair", "tratamiento capilar", 
        "mascarilla capilar", "aceite capilar", "scalp", "hair care"
    ],
    "Ambientadores y Difusores": [
        "ambientador", "ambientadores", "difusor", "difusores", 
        "home fragrance", "velas", "candle", "diffuser"
    ]
}

MAPEO_ML = {
    "GTIN_EAN": ["Código universal de producto"],
    "SKU_Variante": ["SKU"],
    "Stock": ["Stock"],
    "Marca": ["Marca"],
    "Imagenes_URL": ["Fotos"],
    "Descripcion_Parrafo": ["Descripción"],
    "Ancho_cm": ["Ancho (cm)"],
    "Alto_cm": ["Alto (cm)"],
    "Largo_cm": ["Profundidad (cm)"],
    "Apto_Para_Piel": ["Tipo de piel"],
    "Vegano": ["Vegano"],
    "Cruelty_Free": ["Libre de crueldad"],
    "Ingredientes_Completos_INCI": ["Ingredientes"]
}

app = FastAPI(title="Kosmeticos ETL API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------------------------------------------------------
# UTILIDADES Y PROCESAMIENTO DE DICCIONARIOS
# ---------------------------------------------------------

def _normalizar_header(valor: Any) -> str:
    if valor is None:
        return ""
    return re.sub(r"\s+", " ", str(valor)).strip().upper()

def _rango_fecha_oferta_aleatorio() -> Tuple[str, str]:
    inicio = date.today() + timedelta(days=random.randint(0, 3))
    fin = inicio + timedelta(days=random.randint(60, 120))
    return inicio.isoformat(), fin.isoformat()

def _leer_un_diccionario(path: str) -> pd.DataFrame:
    """Optimización: Se usa read_only=True en openpyxl para evitar uso masivo de RAM."""
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.worksheets[0]

    fila_header = None
    encabezados = {}
    
    # Búsqueda de encabezados dentro de las primeras 10 filas
    for r_idx, row in enumerate(ws.iter_rows(max_row=10, values_only=True), start=1):
        valores = {c_idx: _normalizar_header(val) for c_idx, val in enumerate(row, start=1)}
        if "SKU" in valores.values():
            fila_header = r_idx
            encabezados = valores
            break
            
    if fila_header is None:
        wb.close()
        return pd.DataFrame()

    col_sku = next((c for c, v in encabezados.items() if v == "SKU"), None)
    col_isp = next((c for c, v in encabezados.items() if v == "ISP"), None)
    col_web = next((c for c, v in encabezados.items() if "WEB" in v and "PRICE" in v), None)
    col_estipulado = next((c for c, v in encabezados.items() if "FALABELLA" in v and "RIPLEY" in v), None)

    if col_sku is None:
        wb.close()
        return pd.DataFrame()

    filas = []
    for r_idx, row in enumerate(ws.iter_rows(min_row=fila_header + 1, values_only=True), start=fila_header + 1):
        sku = row[col_sku - 1] if col_sku <= len(row) else None
        if sku is None or str(sku).strip() == "":
            continue
            
        filas.append({
            "SKU": str(sku).strip(),
            "ISP": row[col_isp - 1] if col_isp and col_isp <= len(row) else None,
            "Precio_Web": row[col_web - 1] if col_web and col_web <= len(row) else None,
            "Precio_Estipulado": row[col_estipulado - 1] if col_estipulado and col_estipulado <= len(row) else None,
        })
    wb.close()
    return pd.DataFrame(filas)

def obtener_datos_consolidados_diccionarios() -> pd.DataFrame:
    archivos = [f for f in os.listdir(DICCIONARIOS_DIR) if f.endswith(('.xlsx', '.xls')) and not f.startswith('~')]
    if not archivos:
        return pd.DataFrame()

    dfs = []
    for archivo in archivos:
        path = os.path.join(DICCIONARIOS_DIR, archivo)
        try:
            df = _leer_un_diccionario(path)
            if not df.empty:
                dfs.append(df)
        except Exception as e:
            logger.error(f"Error leyendo el diccionario {archivo}: {e}")

    if not dfs:
        return pd.DataFrame()

    diccionario_maestro = pd.concat(dfs, ignore_index=True)
    return diccionario_maestro.drop_duplicates(subset=['SKU'], keep='last')

def cruzar_main_con_diccionario(df_main: pd.DataFrame) -> pd.DataFrame:
    df_diccionario = obtener_datos_consolidados_diccionarios()

    if 'SKU_Variante' in df_main.columns:
        df_main['SKU_Variante'] = df_main['SKU_Variante'].astype(str).str.strip()

    if df_diccionario.empty:
        df_cruzado = df_main.copy()
    else:
        df_cruzado = pd.merge(
            df_main,
            df_diccionario,
            left_on='SKU_Variante',
            right_on='SKU',
            how='left'
        )

    if 'Precio_Estipulado' in df_cruzado.columns:
        df_cruzado['Precio_Base_Final'] = df_cruzado['Precio_Estipulado'].combine_first(df_cruzado.get('Precio_Regular'))
    else:
        df_cruzado['Precio_Base_Final'] = df_cruzado.get('Precio_Regular')

    if 'Precio_Web' in df_cruzado.columns:
        df_cruzado['Precio_Oferta_Final'] = df_cruzado['Precio_Web'].combine_first(df_cruzado.get('Precio_Oferta'))
    else:
        df_cruzado['Precio_Oferta_Final'] = df_cruzado.get('Precio_Oferta')

    return df_cruzado

# ---------------------------------------------------------
# LÓGICA DE CANALES (MERCADO LIBRE, FALABELLA, RIPLEY, WOOCOMMERCE)
# ---------------------------------------------------------

def _mapear_a_hoja_existente(nombre_categoria_detectada: str, lista_hojas: List[str]) -> str:
    for hoja in lista_hojas:
        if nombre_categoria_detectada.lower() in hoja.lower() or hoja.lower() in nombre_categoria_detectada.lower():
            return hoja
            
    palabras_clave_hoja = {
        "Cuidado Facial": ["facial", "rostro", "skincare", "piel"],
        "Maquillaje": ["maquillaje", "makeup"],
        "Cuidado Corporal": ["corporal", "cuerpo", "body"],
        "Cuidado Capilar": ["capilar", "cabello", "hair"]
    }
    
    keywords_familia = palabras_clave_hoja.get(nombre_categoria_detectada, [])
    for hoja in lista_hojas:
        if any(kw in hoja.lower() for kw in keywords_familia):
            return hoja

    for hoja in lista_hojas:
        if "otro" in hoja.lower():
            return hoja

    return lista_hojas[0]

def _elegir_hoja_ml(fila: pd.Series, hojas_disponibles: List[str]) -> str:
    texto_completo = f"{fila.get('Tipo_Producto', '')} {fila.get('Categorias', '')} {fila.get('Titulo_MercadoLibre', '')} {fila.get('Nombre_General', '')}".lower()

    for cat_familia, keywords in MAPEO_CATEGORIAS_ML.items():
        if any(re.search(r'\b' + re.escape(kw) + r'\b', texto_completo) for kw in keywords):
            return _mapear_a_hoja_existente(cat_familia, hojas_disponibles)

    return _mapear_a_hoja_existente("Otros", hojas_disponibles)

def procesar_mercado_libre(df_datos: pd.DataFrame, archivo: str) -> str:
    dir_ml = os.path.join(PLANTILLAS_DIR, "mercado libre")
    archivos_ml = [f for f in os.listdir(dir_ml) if f.endswith('.xlsx') and not f.startswith('~')]
    if not archivos_ml:
        raise HTTPException(status_code=400, detail="No se encontró ninguna plantilla en Plantillas/mercado libre")

    wb = load_workbook(os.path.join(dir_ml, archivos_ml[0]))
    hojas_categoria = list(wb.sheetnames)
    
    filas_por_hoja: Dict[str, List[pd.Series]] = {}
    for _, fila in df_datos.iterrows():
        hoja = _elegir_hoja_ml(fila, hojas_categoria)
        filas_por_hoja.setdefault(hoja, []).append(fila)

    for hoja, filas in filas_por_hoja.items():
        ws = wb[hoja]
        col_indices = {}
        
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=3, column=col).value
            if not val:
                continue
            val = str(val).strip()
            if val.startswith("Título"):
                col_indices.setdefault("Titulo_MercadoLibre", col)
            elif val == "Precio":
                col_indices.setdefault("Precio_Base_Final", col)
            else:
                for col_main, variantes in MAPEO_ML.items():
                    if val in variantes:
                        col_indices.setdefault(col_main, col)

        # Determinar primera fila libre
        max_row_combinada = 3
        for rango in ws.merged_cells.ranges:
            if rango.max_row > max_row_combinada:
                max_row_combinada = rango.max_row
        fila_actual = max_row_combinada + 1

        for fila in filas:
            for col_main, col_dest in col_indices.items():
                valor = fila.get(col_main)
                if col_main == "Peso_g" and pd.notna(valor):
                    valor = round(float(valor) / 1000, 3)
                if pd.notna(valor):
                    ws.cell(row=fila_actual, column=col_dest).value = valor
            fila_actual += 1

    ruta_salida = os.path.join(DB_DIR, f"MercadoLibre_Listo_{archivo}")
    wb.save(ruta_salida)
    return ruta_salida

def procesar_falabella(df_datos: pd.DataFrame, archivo: str) -> str:
    dir_falabella = os.path.join(PLANTILLAS_DIR, "falabella")
    archivos_fala = [f for f in os.listdir(dir_falabella) if f.endswith('.xlsx') and not f.startswith('~')]
    if not archivos_fala:
        raise HTTPException(status_code=400, detail="No se encontró ninguna plantilla en Plantillas/falabella")

    wb = load_workbook(os.path.join(dir_falabella, archivos_fala[0]))
    ws = wb['Subir plantilla']
    
    mapeo = {
        'Nombre_General': 'Nombre #39', 'Marca': 'Marca #26', 'Descripcion_Parrafo': 'Descripción #53',
        'SKU_Variante': 'SKU del vendedor #29', 'GTIN_EAN': 'Código de barras #56',
        'Stock': 'QuantityFalabella #25', 'Ingredientes_Completos_INCI': 'Ingredientes #36104',
        'Sugerencia_de_Uso': 'InstruccionesDeUso #36103', 'Apto_Para_Piel': 'TipoDePiel #391',
        'Contenido': 'MedidaVolumen #405', 'Peso_g': 'Peso del paquete #8',
        'Largo_cm': 'Largo del paquete #33', 'Ancho_cm': 'Ancho del paquete #60',
        'Alto_cm': 'Alto del paquete #47', 'SKU_Padre': 'SKU Padre #3',
        'Precio_Base_Final': 'PriceFalabella #52', 'Precio_Oferta_Final': 'SalePriceFalabella #18',
    }
    if 'ISP' in df_datos.columns: mapeo['ISP'] = 'ResolucionIsp #402'

    col_indices, col_fecha_inicio, col_fecha_fin = {}, None, None
    for col in range(1, ws.max_column + 1):
        val = str(ws.cell(row=4, column=col).value or '').strip()
        if val == 'SaleStartDateFalabella #45': col_fecha_inicio = col
        elif val == 'SaleEndDateFalabella #31': col_fecha_fin = col
        else:
            for col_main, col_fala in mapeo.items():
                if val == col_fala: col_indices[col_main] = col

    fecha_inicio, fecha_fin = _rango_fecha_oferta_aleatorio()
    fila_inicio = 5
    for _, fila in df_datos.iterrows():
        for col_main, col_dest in col_indices.items():
            if col_main in fila and pd.notna(fila[col_main]):
                ws.cell(row=fila_inicio, column=col_dest).value = fila[col_main]
        if pd.notna(fila.get('Precio_Oferta_Final')):
            if col_fecha_inicio: ws.cell(row=fila_inicio, column=col_fecha_inicio).value = fecha_inicio
            if col_fecha_fin: ws.cell(row=fila_inicio, column=col_fecha_fin).value = fecha_fin
        fila_inicio += 1

    ruta_salida = os.path.join(DB_DIR, f"Falabella_Listo_{archivo}")
    wb.save(ruta_salida)
    return ruta_salida

def procesar_ripley(df_datos: pd.DataFrame, archivo: str) -> str:
    dir_ripley = os.path.join(PLANTILLAS_DIR, "ripley")
    archivos_ripley = [f for f in os.listdir(dir_ripley) if f.endswith('.xlsx') and not f.startswith('~')]
    if not archivos_ripley:
        raise HTTPException(status_code=400, detail="No se encontró ninguna plantilla en Plantillas/ripley")

    wb = load_workbook(os.path.join(dir_ripley, archivos_ripley[0]))
    ws = wb['Data']
    
    mapeo = {
        'SKU_Variante': ['sku_seller', 'SKU de oferta', 'ID de producto'],
        'Titulo_MercadoLibre': 'Titulo', 'Descripcion_Parrafo': 'Descripcion',
        'Marca': 'Marca', 'Tipo_Producto': 'Tipo de Producto',
        'Contenido': ['Contenido', 'Contenido Producto'], 'Apto_Para_Piel': 'Tipo de piel',
        'Nombre_General': 'Nombre del Producto', 'Ingredientes_Principales': 'Activo Cosmético',
        'Sugerencia_de_Uso': 'Modo de Empleo', 'Ingredientes_Completos_INCI': 'Ingredientes',
        'Precio_Base_Final': 'Precio de la oferta', 'Precio_Oferta_Final': 'Precio de descuento',
    }
    if 'ISP' in df_datos.columns: mapeo['ISP'] = 'N° de Registro ISP'

    col_indices, col_fecha_inicio, col_fecha_fin = {}, None, None
    for col in range(1, ws.max_column + 1):
        val = str(ws.cell(row=1, column=col).value or '').strip()
        if val == 'Fecha de inicio del descuento': col_fecha_inicio = col
        elif val == 'Fecha de finalización del descuento': col_fecha_fin = col
        else:
            for col_main, col_rips in mapeo.items():
                targets = col_rips if isinstance(col_rips, list) else [col_rips]
                if val in targets:
                    col_indices.setdefault(col_main, []).append(col)

    fecha_inicio, fecha_fin = _rango_fecha_oferta_aleatorio()
    fila_inicio = 2
    for _, fila in df_datos.iterrows():
        for col_main, columnas_dest in col_indices.items():
            if col_main in fila and pd.notna(fila[col_main]):
                for col_dest in columnas_dest:
                    ws.cell(row=fila_inicio, column=col_dest).value = fila[col_main]
        if pd.notna(fila.get('Precio_Oferta_Final')):
            if col_fecha_inicio: ws.cell(row=fila_inicio, column=col_fecha_inicio).value = fecha_inicio
            if col_fecha_fin: ws.cell(row=fila_inicio, column=col_fecha_fin).value = fecha_fin
        fila_inicio += 1

    ruta_salida = os.path.join(DB_DIR, f"Ripley_Listo_{archivo}")
    wb.save(ruta_salida)
    return ruta_salida

def procesar_woocommerce(df_datos: pd.DataFrame, archivo: str) -> str:
    df_wp = pd.DataFrame()
    df_wp['SKU'] = df_datos['SKU_Variante'] if 'SKU_Variante' in df_datos else ""
    df_wp['Name'] = df_datos['Nombre_General'] if 'Nombre_General' in df_datos else ""
    df_wp['Short description'] = df_datos['Descripcion_BulletPoints'] if 'Descripcion_BulletPoints' in df_datos else ""
    df_wp['Description'] = df_datos['Descripcion_Parrafo'] if 'Descripcion_Parrafo' in df_datos else ""
    df_wp['Regular price'] = df_datos['Precio_Oferta_Final'] if 'Precio_Oferta_Final' in df_datos else ""
    
    nombre_base = os.path.splitext(archivo)[0]
    ruta_salida = os.path.join(DB_DIR, f"WooCommerce_Listo_{nombre_base}.csv")
    df_wp.to_csv(ruta_salida, index=False, encoding="utf-8-sig")
    return ruta_salida

# ---------------------------------------------------------
# ENDPOINTS API
# ---------------------------------------------------------

app.mount("/static", StaticFiles(directory=WEB_DIR), name="static")

@app.get("/", response_class=HTMLResponse)
async def serve_frontend():
    index_path = os.path.join(WEB_DIR, "index.html")
    if not os.path.exists(index_path):
        raise HTTPException(status_code=404, detail="index.html no encontrado")
    with open(index_path, "r", encoding="utf-8") as f:
        return f.read()

@app.get("/api/marketing/download-template/")
async def download_template():
    file_path = os.path.join(PLANTILLAS_DIR, "plantilla madre.xlsx")
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Plantilla madre no encontrada")
    return FileResponse(path=file_path, filename="plantilla madre.xlsx")

@app.post("/api/marketing/upload-excel/")
async def upload_main_excel(archivo: UploadFile = File(...)):
    if not archivo.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="El archivo debe ser un Excel (.xlsx o .xls).")

    file_location = os.path.join(DB_DIR, archivo.filename)
    with open(file_location, "wb") as buffer:
        shutil.copyfileobj(archivo.file, buffer)

    try:
        df = pd.read_excel(file_location, sheet_name=0)
        df.columns = df.columns.str.strip()
    except Exception as e:
        os.remove(file_location)
        raise HTTPException(status_code=400, detail=f"No se pudo leer el Excel: {e}")

    faltantes = [c for c in COLUMNAS_OBLIGATORIAS_MAIN if c not in df.columns]
    if faltantes:
        os.remove(file_location)
        return JSONResponse(
            status_code=status.HTTP_400_BAD_REQUEST,
            content={"error": "Estructura inválida", "columnas_faltantes": faltantes}
        )

    df_clean = df.dropna(how="all")
    if df_clean.empty:
        os.remove(file_location)
        raise HTTPException(status_code=400, detail="El archivo no tiene filas de datos validas.")

    return {"mensaje": "Archivo validado y guardado", "archivo": archivo.filename, "filas": len(df_clean)}

@app.get("/api/dev/archivos-disponibles/")
async def listar_archivos():
    return {"archivos": [f for f in os.listdir(DB_DIR) if f.endswith(('.xlsx', '.xls', '.csv'))]}

@app.post("/api/dev/upload-diccionario/")
async def upload_diccionario(archivo: UploadFile = File(...)):
    if not archivo.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="El archivo debe ser un Excel (.xlsx o .xls).")

    file_location = os.path.join(DICCIONARIOS_DIR, archivo.filename)
    with open(file_location, "wb") as buffer:
        shutil.copyfileobj(archivo.file, buffer)

    try:
        df = _leer_un_diccionario(file_location)
    except Exception as e:
        if os.path.exists(file_location): os.remove(file_location)
        raise HTTPException(status_code=400, detail=f"Error leyendo el diccionario: {e}")

    if df.empty:
        os.remove(file_location)
        raise HTTPException(status_code=400, detail="No se encontró columna 'SKU' o no hay datos.")

    # Limpiar diccionarios antiguos
    for f in os.listdir(DICCIONARIOS_DIR):
        if f != archivo.filename:
            try:
                os.remove(os.path.join(DICCIONARIOS_DIR, f))
            except OSError:
                pass

    return {"mensaje": "Diccionario actualizado correctamente", "archivo": archivo.filename, "productos_leidos": len(df)}

@app.get("/api/dev/diccionario-actual/")
async def diccionario_actual():
    archivos = [f for f in os.listdir(DICCIONARIOS_DIR) if f.endswith(('.xlsx', '.xls')) and not f.startswith('~')]
    return {"archivo": archivos[0] if archivos else None}

@app.delete("/api/dev/eliminar-diccionario/")
async def eliminar_diccionario():
    eliminados = 0
    for f in os.listdir(DICCIONARIOS_DIR):
        if f.endswith(('.xlsx', '.xls')) and not f.startswith('~'):
            os.remove(os.path.join(DICCIONARIOS_DIR, f))
            eliminados += 1
    return {"mensaje": "Diccionario eliminado", "archivos_eliminados": eliminados}

@app.get("/api/dev/procesar-tienda/{tienda_id}")
async def procesar_tienda(tienda_id: str, archivo: str = Query(...)):
    file_path = os.path.join(DB_DIR, archivo)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="El archivo no existe en la BD")
    
    try:
        df_main = pd.read_excel(file_path)
        df_datos = cruzar_main_con_diccionario(df_main)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error al leer Plantilla Madre: {e}")

    processors = {
        "ml": procesar_mercado_libre,
        "mercadolibre": procesar_mercado_libre,
        "falabella": procesar_falabella,
        "ripley": procesar_ripley,
        "woocommerce": procesar_woocommerce,
    }

    processor = processors.get(tienda_id.lower())
    if not processor:
        raise HTTPException(status_code=400, detail=f"Canal '{tienda_id}' no soportado.")

    try:
        ruta_salida = processor(df_datos, archivo)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error procesando plantilla {tienda_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Error en motor de inyección: {e}")

    return FileResponse(path=ruta_salida, filename=os.path.basename(ruta_salida))

# Endpoints unificados sin duplicar código con meta-programación
@app.api_route("/export/{canal}", methods=["GET", "POST"])
async def export_alias(canal: str, archivo: str = Query(...)):
    canal_map = {"mercadolibre": "ml", "falabella": "falabella", "ripley": "ripley", "woocommerce": "woocommerce"}
    tienda_id = canal_map.get(canal.lower())
    if not tienda_id:
        raise HTTPException(status_code=400, detail="Canal de exportación no válido")
    return await procesar_tienda(tienda_id, archivo)

if __name__ == "__main__":
    puerto = int(os.environ.get("PORT", 8000))
    solo_local = puerto == 8000 and "PORT" not in os.environ
    logger.info("Iniciando Kosmeticos ETL...")
    if solo_local:
        webbrowser.open("http://127.0.0.1:8000")
    uvicorn.run(app, host="0.0.0.0", port=puerto)